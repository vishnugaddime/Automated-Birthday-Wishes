import openpyxl
from datetime import date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import sys

import nexmo

from fbchat import Client
from fbchat.models import *


def send_mail(e):
    email = "vishnugaddime00@gmail.com"
    send_to_email = e
    password = "Vaishnavi@123"
    subject = 'Wishes'

    msg = MIMEMultipart('alternative')
    msg['From'] = email
    msg['To'] = send_to_email
    msg['Subject'] = subject

    # Attach the message to the MIMEMultipart object
    msg.attach(MIMEText('<html><body><pre><center><h1 style="font-family: "sans-serif"; font-color: "#18D9ECF"; font-size: 20px;"><b><i>"Another year, another challenge my friend. But, you need to always remember to<br> keep smiling and have a positive outlook on life,<br> and everything is going to be fine.<br>Another birthday means your life journey is incomplete,<br> may your path be paved with success and <br> guided by love. Best wishes, my friend. HAPPY BIRTHDAY !"</i></b></h1></pre></center></body></html>', 'html'))

    filename = "wishes.jpg"
    attachment = open(".\wishes.jpg", "rb")

    # instance of MIMEBase and named as p
    p = MIMEBase('application', 'octet-stream')

    # To change the payload into encoded form
    p.set_payload((attachment).read(0))

    # encode into base64
    encoders.encode_base64(p)

    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

    # attach the instance 'p' to instance 'msg'
    msg.attach(p)

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(email, password)
    text = msg.as_string() # You now need to convert the MIMEMultipart object to a string to send
    server.sendmail(email, send_to_email, text)
    print ("MISSION SUCCESSFULL !")
    server.quit()

def whatsapp(name):
    # Enter the  path to chromedriver
    driver = webdriver.Chrome('.\chromedriver')

    driver.get("https://web.whatsapp.com/")
    wait = WebDriverWait(driver, 600)

    # Enter the Name or Group
    target = name

    # Enter the message Here
    string = "Another year, another challenge my friend. But, you need to always remember to keep smiling and have a positive outlook on life,and everything is going to be fine. Another birthday means your life journey is incomplete,may your path be paved with success and guided by love. Best wishes, my friend.\n Wish You HAPPY BIRTHDAY !"

    x_arg = '//span[contains(@title,' + target + ')]'
    group_title = wait.until(EC.presence_of_element_located((
        By.XPATH, x_arg)))
    group_title.click()

    message = driver.find_elements_by_xpath('//*[@id="main"]/footer/div[1]/div[2]/div/div[2]')[0]

    message.send_keys(string)

    sendbutton = driver.find_elements_by_xpath('//*[@id="main"]/footer/div[1]/div[3]/button')[0]
    sendbutton.click()

    print("Whats App Msg is send")

    driver.close()


def sms(number) :

    client = nexmo.Client(key='b54922c4', secret='hdXQg3MZb2mwS5jh')

    client.send_message({
        'from': 'Nexmo',
        'to': number,
        'text': 'Another year, another challenge my friend. But, you need to always remember to keep smiling and have a positive outlook on life, WIsh You HAPPY BIRTHDAY !! ',
    })

    print("sms is sent successfully !!")

def fb(f):

    username = "77588993889"
    password = "Vishnu@123"

    client = Client(username, password)

    if client.isLoggedIn():

            # person or group details
            name = f
            friends = client.searchForUsers(name)  # return a list of names
            friend = friends[0]
            msg = " Another year, another challenge my friend. But, you need to always remember to keep smiling and have a positive outlook on life, WIsh You HAPPY BIRTHDAY !!"

            client.send(Message(text=msg), thread_id=str(friend.uid), thread_type=ThreadType.USER)

            client.logout()
            print("Message is sent on fb")

    else:
            print('not logged in')


today = date.today()
day = today.day
month = today.month

wb_obj = openpyxl.load_workbook("test.xlsx")
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=3)
    d = cell_obj.value
    cell_obj = sheet_obj.cell(row=i, column=4)
    m = cell_obj.value
    cell_obj = sheet_obj.cell(row=i, column=2)
    e = cell_obj.value
    cell_obj = sheet_obj.cell(row=i, column=1)
    name = cell_obj.value
    cell_obj = sheet_obj.cell(row=i, column=6)
    number = cell_obj.value
    cell_obj = sheet_obj.cell(row=i, column=7)
    f = cell_obj.value


    if day == d and month == m:
        name = '"'+ name +'"'
       # send_mail(e)
        whatsapp(name)
       # sms(number)
       # fb(f)


